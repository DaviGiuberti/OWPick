import re
import html
import sys
import os
from openpyxl import Workbook

MAP_FILE = "map.txt"
WINRATE_DIR = "winratemaps"
OUTPUT_XLSX = "winrate.xlsx"

pattern = re.compile(
    r'{"name"\s*:\s*"(?P<name>[^"]+)"[^}]*?"winrate"\s*:\s*(?P<wr>\d+(?:\.\d+)?)',
    re.IGNORECASE | re.UNICODE
)

pickrate_search = re.compile(r'"pickrate"\s*:\s*(?P<pr>\d+(?:\.\d+)?)', re.IGNORECASE)

def get_input_files():
    try:
        with open(MAP_FILE, "r", encoding="utf-8") as f:
            map_name = f.read().strip()
    except FileNotFoundError:
        print(f"Erro: '{MAP_FILE}' não encontrado. Escolha um mapa")
        return None, None

    if not map_name:
        print("Erro: 'map.txt' está vazio. Escolha um mapa")
        return None, None

    html_master = os.path.join(WINRATE_DIR, f"{map_name}_Master.html")
    html_grandmaster = os.path.join(WINRATE_DIR, f"{map_name}_Grandmaster.html")

    return html_master, html_grandmaster

def parse_file(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = f.read()
    except FileNotFoundError:
        print(f"Erro: '{path}' não encontrado. Atualize as winrates")
        return None, None, None, None

    text = html.unescape(raw)
    num_map = {}       # winrate numeric
    str_map = {}       # winrate string (comma decimal)
    pick_map = {}      # pickrate numeric
    pick_str_map = {}  # pickrate string (comma decimal)

    for m in pattern.finditer(text):
        name = m.group("name").strip()
        name = name.replace(":", "").replace(".", "")

        wr_text = m.group("wr").strip()
        try:
            wr_num = float(wr_text)
        except ValueError:
            continue

        # procura pickrate dentro do mesmo trecho capturado (se houver)
        snippet = m.group(0)
        pick_m = pickrate_search.search(snippet)
        if pick_m:
            try:
                pr_num = float(pick_m.group("pr"))
            except ValueError:
                pr_num = None
        else:
            pr_num = None

        num_map[name] = wr_num
        str_map[name] = wr_text.replace(".", ",")

        if pr_num is not None:
            pick_map[name] = pr_num
            pick_str_map[name] = f"{pr_num:.2f}".replace(".", ",")

    return num_map, str_map, pick_map, pick_str_map

def executar():
    INPUT_HTML1, INPUT_HTML2 = get_input_files()

    if not INPUT_HTML1 or not INPUT_HTML2:
        return 

    num1, str1, pick1, pickstr1 = parse_file(INPUT_HTML1)
    num2, str2, pick2, pickstr2 = parse_file(INPUT_HTML2)

    if num1 is None or num2 is None:
        return

    names_sorted = sorted(
        set(num1.keys()) | set(num2.keys()) | set(pick1.keys()) | set(pick2.keys()),
        key=lambda s: s.lower()
    )

    if not names_sorted:
        print("Nenhuma winrate encontrada nos arquivos. Atualize as winrates.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Winrates"

    ws["A1"] = "Hero"
    ws["B1"] = "Winrate Master"
    ws["C1"] = "Winrate Grandmaster"
    ws["D1"] = "Average"
    ws["E1"] = "Saida (0,2*avg - 10) * 2"
    ws["F1"] = "Pickrate Score (4%=-2,20%=2)"

    row = 2
    for name in names_sorted:
        ws.cell(row=row, column=1, value=name)

        if name in str1:
            ws.cell(row=row, column=2, value=str1[name])

        if name in str2:
            ws.cell(row=row, column=3, value=str2[name])

        if name in num1 and name in num2:
            avg = (num1[name] + num2[name]) / 2
            ws.cell(row=row, column=4, value=f"{avg:.2f}".replace(".", ","))

            saida = (0.2 * avg - 10.0) * 2
            ws.cell(row=row, column=5, value=f"{saida:.2f}".replace(".", ","))

        # cálculo da média das pickrates e conversão para a escala -2..2
        avg_pick = None
        if name in pick1 and name in pick2:
            avg_pick = (pick1[name] + pick2[name]) / 2
        elif name in pick1:
            avg_pick = pick1[name]
        elif name in pick2:
            avg_pick = pick2[name]

        if avg_pick is not None:
            # escala linear: 4% -> -2 ; 20% -> 2  => score = 0.25 * pick_pct - 3
            score = 0.25 * avg_pick - 3.0
            ws.cell(row=row, column=6, value=f"{score:.2f}".replace(".", ","))

        row += 1

    wb.save(OUTPUT_XLSX)
    print(f"OK — {len(names_sorted)} heróis salvos em '{OUTPUT_XLSX}'.")

if __name__ == "__main__":
    executar()
